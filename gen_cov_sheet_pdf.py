import os
import streamlit as st
import pandas as pd
import pg8000
from io import BytesIO
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from fpdf import FPDF

# Function to establish a database connection
def get_database_connection():
    db_connection = pg8000.connect(
        database=os.environ["SUPABASE_DB_NAME"],
        user=os.environ["SUPABASE_USER"],
        password=os.environ["SUPABASE_PASSWORD"],
        host=os.environ["SUPABASE_HOST"],
        port=os.environ["SUPABASE_PORT"]
    )
    return db_connection

# Function to query the database and fetch data
def fetch_student_data(student_list):
    db_connection = get_database_connection()
    db_cursor = db_connection.cursor()
    
    student_list_string = ', '.join(map(str, student_list))

    db_query = f"""SELECT student_list.name,                  
                    student_list.iatc_id,
                    student_list.nat_id,
                    student_list.class,
                    exam_list.exam_long AS subject,
                    exam_results.score,
                    exam_results.result,
                    exam_results.date
                    FROM exam_results 
                    JOIN student_list ON exam_results.nat_id = student_list.nat_id
                    JOIN exam_list ON exam_results.exam = exam_list.exam
                    WHERE student_list.iatc_id IN ({student_list_string}) AND exam_results.score_index = 1
                    ORDER BY exam_list.srt_exam ASC
                """
    db_cursor.execute(db_query)
    output_data = db_cursor.fetchall()
    db_cursor.close()
    db_connection.close()

    col_names = ['Name', 'IATC ID', 'National ID', 'Class', 'Subject', 'Score', 'Result', 'Date']
    df = pd.DataFrame(output_data, columns=col_names)
    return df

# Function to generate an Excel sheet and return it as a buffer
def create_excel_sheet(student_data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Coversheet"

    # Styling variables
    header_font = Font(bold=True)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Populate static text in specific cells
    sheet["B2"] = "Student Name:"
    sheet["B3"] = "Student IATC ID:"
    sheet["B4"] = "Student National ID:"
    sheet["B5"] = "Student Class:"

    sheet["B2"].font = sheet["B3"].font = sheet["B4"].font = sheet["B5"].font = header_font
    sheet["B2"].alignment = sheet["B3"].alignment = sheet["B4"].alignment = sheet["B5"].alignment = cell_alignment

    sheet["C2"] = student_data['Name'].iloc[0]
    sheet["C3"] = student_data['IATC ID'].iloc[0]
    sheet["C4"] = student_data['National ID'].iloc[0]
    sheet["C5"] = student_data['Class'].iloc[0]

    # Populate table headers
    headers = ['Subject', 'Score', 'Result', 'Date']
    for col_num, header in enumerate(headers, start=2):
        cell = sheet.cell(row=7, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = cell_alignment
        cell.border = thin_border

    # Populate the data table
    for row_num, row_data in enumerate(student_data[['Subject', 'Score', 'Result', 'Date']].values, start=8):
        for col_num, value in enumerate(row_data, start=2):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Adjust column widths
    for col in range(2, 6):  # Columns B to E
        sheet.column_dimensions[get_column_letter(col)].width = 20

    # Save the workbook to a buffer
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer

# Function to convert Excel to PDF using FPDF
def excel_to_pdf(excel_buffer, student_id):
    # Load Excel data
    workbook = pd.ExcelFile(excel_buffer)
    sheet_data = workbook.parse(workbook.sheet_names[0])

    # Create a PDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Title
    pdf.set_font("Arial", style="B", size=14)
    pdf.cell(0, 10, f"Coversheet for Student ID: {student_id}", ln=True, align="C")

    # Table header
    pdf.set_font("Arial", style="B", size=10)
    for header in sheet_data.columns:
        pdf.cell(40, 10, str(header), border=1, align="C")
    pdf.ln()

    # Table rows
    pdf.set_font("Arial", size=10)
    for _, row in sheet_data.iterrows():
        for value in row:
            pdf.cell(40, 10, str(value), border=1, align="C")
        pdf.ln()

    # Save PDF to buffer
    pdf_buffer = BytesIO()
    pdf.output(pdf_buffer)
    pdf_buffer.seek(0)

    return pdf_buffer

# Function to generate PDFs and save them to a zip file
def generate_coversheets_zip(student_list):
    df = fetch_student_data(student_list)

    # Create an in-memory ZIP file
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        for student_id in student_list:
            filtered_df = df[df['IATC ID'] == student_id]

            # Generate Excel and then convert to PDF
            excel_buffer = create_excel_sheet(filtered_df)
            pdf_buffer = excel_to_pdf(excel_buffer, student_id)

            # Add PDF to the zip
            pdf_filename = f"{student_id}.pdf"
            zip_file.writestr(pdf_filename, pdf_buffer.getvalue())

    zip_buffer.seek(0)
    return zip_buffer

# Streamlit interface
st.title("Generate Theory Exam Coversheets (PDF)")
st.write("Enter a list of student IDs and download the PDF coversheets containing the highest result for each subject the student has taken.")

student_ids_input = st.text_area("Enter Student IDs separated by commas (e.g., 151596, 156756, 154960):")
st.write("Need help generating a list of IDs? Download the Excel template:")

# Direct link to the Excel file in GitHub
template_url = "https://github.com/Bayr-Harrison/coversheetgenerator/raw/main/Coversheet%20Generator%20Input.xlsx"
st.markdown(f"[Download Excel Template]({template_url})", unsafe_allow_html=True)

if st.button("Generate Coversheets"):
    try:
        student_list = [int(id.strip()) for id in student_ids_input.split(",")]
        st.write("Generating coversheets...")

        # Generate the zip file in memory
        zip_file = generate_coversheets_zip(student_list)

        # Offer the zip file for download
        st.download_button(
            label="Download All Coversheets as ZIP",
            data=zip_file,
            file_name="coversheets.zip",
            mime="application/zip"
        )
        st.success("Coversheets zip generated successfully!")

    except Exception as e:
        st.error(f"An error occurred: {e}")
